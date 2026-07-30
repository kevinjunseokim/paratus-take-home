from __future__ import annotations

import io
import logging
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any, Literal

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.afsc import AfscEngine, get_afsc_engine
from app.db.models import MemberRecord, RosterUploadRecord
from app.db.repositories.roster import RosterRepository
from app.enums import IssueSeverity, PersonnelType, UploadStatus
from app.exceptions import AfscValidationError, RosterImportError, RosterNotFoundError
from app.schemas import IssueOut, PreviewOut, PreviewRowOut, UploadOut

logger = logging.getLogger(__name__)

HEADER_ALIASES: dict[str, set[str]] = {
    "dodid": {"dodid", "dod_id", "edipi", "person_id", "member_id"},
    "display_name": {"name", "display_name", "member", "member_name", "full_name"},
    "first_name": {"first_name", "firstname", "given_name"},
    "last_name": {"last_name", "lastname", "surname", "family_name"},
    "rank": {"rank", "grade", "pay_grade"},
    "afsc": {"afsc", "dafsc", "pafsc", "specialty_code"},
    "personnel_type": {"personnel_type", "type", "category", "officer_enlisted"},
}


@dataclass
class _AcceptedMember:
    dodid: str
    display_name: str
    rank: str | None
    personnel_type: PersonnelType | None
    afsc: str
    normalized_afsc: str
    source_row_number: int


@dataclass
class _RowResult:
    row: int | None
    status: Literal["success", "failure"]
    dodid: str | None = None
    display_name: str | None = None
    afsc: str | None = None
    normalized_afsc: str | None = None
    personnel_type: PersonnelType | None = None
    reason: str | None = None
    severity: IssueSeverity | None = None


class ImportService:
    def __init__(
        self,
        session: Session,
        afsc_engine: AfscEngine | None = None,
        *,
        max_upload_bytes: int = 5 * 1024 * 1024,
    ) -> None:
        self._session = session
        self._repo = RosterRepository(session)
        self._afsc = afsc_engine or get_afsc_engine()
        self._max_upload_bytes = max_upload_bytes

    def preview_workbook(self, *, filename: str, content: bytes) -> PreviewOut:
        upload = self._create_upload(filename=filename, content=content)
        try:
            preview = self._process_upload(upload, content, activate=False)
            self._session.commit()
            return preview
        except RosterImportError:
            self._session.rollback()
            raise
        except Exception:
            self._session.rollback()
            logger.exception("Unexpected failure previewing workbook %s", filename)
            raise

    def commit_upload(self, upload_id: uuid.UUID) -> UploadOut:
        upload = self._repo.get(upload_id)
        if upload is None:
            raise RosterNotFoundError("Upload not found")
        if upload.is_active:
            return self._to_upload_out(upload)
        if upload.status != UploadStatus.PENDING:
            raise RosterImportError("Only pending previews can be committed")
        if upload.accepted_rows <= 0:
            raise RosterImportError("Preview has no accepted rows to commit")
        if not self._repo.list_members_for_upload(upload.id):
            raise RosterImportError("Preview has no staged members to commit")

        self._repo.activate_succeeded(
            upload,
            total_rows=upload.total_rows,
            accepted_rows=upload.accepted_rows,
            rejected_rows=upload.rejected_rows,
        )
        self._session.commit()
        loaded = self._repo.get(upload.id)
        if loaded is None:
            raise RosterImportError("Upload disappeared after activation")
        return self._to_upload_out(loaded)

    def discard_upload(self, upload_id: uuid.UUID) -> None:
        upload = self._repo.get(upload_id)
        if upload is None:
            raise RosterNotFoundError("Upload not found")
        if upload.is_active:
            raise RosterImportError("Cannot discard the active roster")
        self._repo.delete_upload(upload)
        self._session.commit()

    def clear_roster(self) -> int:
        deleted = self._repo.clear_all_members()
        self._session.commit()
        return deleted

    def discard_stale_pending(self, *, max_age: timedelta) -> int:
        cutoff = datetime.now(timezone.utc) - max_age
        stale = []
        for upload in self._repo.list_all():
            if upload.status != UploadStatus.PENDING or upload.is_active:
                continue
            uploaded_at = upload.uploaded_at
            if uploaded_at is None:
                continue
            if uploaded_at.tzinfo is None:
                uploaded_at = uploaded_at.replace(tzinfo=timezone.utc)
            if uploaded_at >= cutoff:
                continue
            stale.append(upload)
        for upload in stale:
            self._repo.delete_upload(upload)
        if stale:
            self._session.commit()
        return len(stale)

    def _create_upload(self, *, filename: str, content: bytes) -> RosterUploadRecord:
        if not filename.lower().endswith((".xlsx", ".xlsm")):
            raise RosterImportError("Only .xlsx or .xlsm workbooks are supported")
        if len(content) > self._max_upload_bytes:
            raise RosterImportError("Upload exceeds maximum allowed size")
        return self._repo.create_pending(
            filename=filename,
            ruleset_version=self._afsc.ruleset_version,
        )

    def _process_upload(
        self,
        upload: RosterUploadRecord,
        content: bytes,
        *,
        activate: bool,
    ) -> PreviewOut:
        successes: list[_RowResult] = []
        failures: list[_RowResult] = []

        try:
            rows = self._parse_rows(content)
        except Exception as exc:  # noqa: BLE001
            self._repo.add_issue(
                upload,
                row_number=None,
                field="file",
                raw_value=upload.filename,
                reason=f"Failed to parse workbook: {exc}",
            )
            self._repo.mark_failed(upload, total_rows=0, accepted_rows=0, rejected_rows=0)
            self._session.flush()
            failures.append(
                _RowResult(
                    row=None,
                    status="failure",
                    reason=f"Failed to parse workbook: {exc}",
                    severity=IssueSeverity.ERROR,
                )
            )
            return self._to_preview_out(upload, successes, failures)

        prior = self._prior_members_by_dodid()
        total = len(rows)
        accepted = 0
        rejected = 0
        seen: set[str] = set()
        overlay: dict[str, _AcceptedMember] = {
            dodid: self._from_record(record) for dodid, record in prior.items()
        }

        for row in rows:
            row_number = row["row_number"]
            dodid = row.get("dodid")
            display_name = row.get("display_name")
            afsc_raw = row.get("afsc")
            rank = row.get("rank")
            type_raw = row.get("personnel_type")

            if not dodid:
                rejected += 1
                reason = "Missing DODID"
                self._repo.add_issue(
                    upload,
                    row_number=row_number,
                    field="dodid",
                    raw_value=None,
                    reason=reason,
                )
                failures.append(
                    _RowResult(
                        row=row_number,
                        status="failure",
                        display_name=display_name,
                        afsc=str(afsc_raw) if afsc_raw else None,
                        reason=reason,
                        severity=IssueSeverity.ERROR,
                    )
                )
                continue

            if dodid in seen:
                rejected += 1
                reason = "Duplicate DODID in this upload"
                self._repo.add_issue(
                    upload,
                    row_number=row_number,
                    field="dodid",
                    raw_value=dodid,
                    reason=reason,
                )
                failures.append(
                    _RowResult(
                        row=row_number,
                        status="failure",
                        dodid=dodid,
                        display_name=display_name,
                        afsc=str(afsc_raw) if afsc_raw else None,
                        reason=reason,
                        severity=IssueSeverity.ERROR,
                    )
                )
                continue
            seen.add(dodid)

            if not display_name and not afsc_raw:
                rejected += 1
                reason = (
                    "Empty member fields; keeping prior roster entry"
                    if dodid in prior
                    else "Empty member fields and no prior roster entry"
                )
                severity = IssueSeverity.WARNING if dodid in prior else IssueSeverity.ERROR
                self._repo.add_issue(
                    upload,
                    row_number=row_number,
                    field="dodid",
                    raw_value=dodid,
                    reason=reason,
                    severity=severity,
                )
                failures.append(
                    _RowResult(
                        row=row_number,
                        status="failure",
                        dodid=dodid,
                        reason=reason,
                        severity=severity,
                    )
                )
                continue

            if not display_name:
                rejected += 1
                reason, severity = self._keep_prior_or_reject(
                    upload,
                    prior=prior,
                    dodid=dodid,
                    row_number=row_number,
                    field="display_name",
                    raw_value=None,
                    reason="Missing member name",
                )
                failures.append(
                    _RowResult(
                        row=row_number,
                        status="failure",
                        dodid=dodid,
                        afsc=str(afsc_raw) if afsc_raw else None,
                        reason=reason,
                        severity=severity,
                    )
                )
                continue

            if not afsc_raw:
                rejected += 1
                reason, severity = self._keep_prior_or_reject(
                    upload,
                    prior=prior,
                    dodid=dodid,
                    row_number=row_number,
                    field="afsc",
                    raw_value=None,
                    reason="Missing AFSC",
                )
                failures.append(
                    _RowResult(
                        row=row_number,
                        status="failure",
                        dodid=dodid,
                        display_name=display_name,
                        reason=reason,
                        severity=severity,
                    )
                )
                continue

            try:
                parsed = self._afsc.validate_member_afsc(str(afsc_raw))
            except AfscValidationError as exc:
                rejected += 1
                reason, severity = self._keep_prior_or_reject(
                    upload,
                    prior=prior,
                    dodid=dodid,
                    row_number=row_number,
                    field="afsc",
                    raw_value=str(afsc_raw),
                    reason=str(exc),
                )
                failures.append(
                    _RowResult(
                        row=row_number,
                        status="failure",
                        dodid=dodid,
                        display_name=display_name,
                        afsc=str(afsc_raw),
                        reason=reason,
                        severity=severity,
                    )
                )
                continue

            personnel_type = parsed.personnel_type
            if type_raw:
                mapped = self._parse_personnel_type(str(type_raw))
                if mapped is None:
                    self._repo.add_issue(
                        upload,
                        row_number=row_number,
                        field="personnel_type",
                        raw_value=str(type_raw),
                        reason="Unrecognized personnel type; using AFSC-derived type",
                        severity=IssueSeverity.WARNING,
                    )
                elif mapped != parsed.personnel_type:
                    rejected += 1
                    reason, severity = self._keep_prior_or_reject(
                        upload,
                        prior=prior,
                        dodid=dodid,
                        row_number=row_number,
                        field="personnel_type",
                        raw_value=str(type_raw),
                        reason=(
                            f"Row says {mapped.value}, but AFSC {parsed.normalized} "
                            f"is {parsed.personnel_type.value}"
                        ),
                    )
                    failures.append(
                        _RowResult(
                            row=row_number,
                            status="failure",
                            dodid=dodid,
                            display_name=display_name,
                            afsc=str(afsc_raw),
                            normalized_afsc=parsed.normalized,
                            personnel_type=parsed.personnel_type,
                            reason=reason,
                            severity=severity,
                        )
                    )
                    continue

            overlay[dodid] = _AcceptedMember(
                dodid=dodid,
                display_name=display_name,
                rank=str(rank).strip() if rank else None,
                personnel_type=personnel_type,
                afsc=str(afsc_raw).strip(),
                normalized_afsc=parsed.normalized,
                source_row_number=row_number,
            )
            accepted += 1
            successes.append(
                _RowResult(
                    row=row_number,
                    status="success",
                    dodid=dodid,
                    display_name=display_name,
                    afsc=str(afsc_raw).strip(),
                    normalized_afsc=parsed.normalized,
                    personnel_type=personnel_type,
                )
            )

        if accepted == 0:
            reason = (
                "No valid member rows; prior active roster left unchanged"
                if prior
                else "No valid member rows to activate"
            )
            self._repo.add_issue(
                upload,
                row_number=None,
                field="file",
                raw_value=upload.filename,
                reason=reason,
            )
            self._repo.mark_failed(
                upload,
                total_rows=total,
                accepted_rows=0,
                rejected_rows=rejected,
            )
            self._session.flush()
            return self._to_preview_out(upload, successes, failures)

        for member in overlay.values():
            self._repo.add_member(
                upload,
                dodid=member.dodid,
                display_name=member.display_name,
                rank=member.rank,
                personnel_type=member.personnel_type,
                afsc=member.afsc,
                normalized_afsc=member.normalized_afsc,
                source_row_number=member.source_row_number,
            )

        if activate:
            self._repo.activate_succeeded(
                upload,
                total_rows=total,
                accepted_rows=accepted,
                rejected_rows=rejected,
            )
        else:
            self._repo.stage_pending(
                upload,
                total_rows=total,
                accepted_rows=accepted,
                rejected_rows=rejected,
            )
        self._session.flush()
        loaded = self._repo.get(upload.id)
        if loaded is None:
            raise RosterImportError("Upload disappeared after staging")
        return self._to_preview_out(loaded, successes, failures)

    def _keep_prior_or_reject(
        self,
        upload: RosterUploadRecord,
        *,
        prior: dict[str, MemberRecord],
        dodid: str,
        row_number: int,
        field: str,
        raw_value: str | None,
        reason: str,
    ) -> tuple[str, IssueSeverity]:
        if dodid in prior:
            full_reason = f"{reason}; keeping prior roster entry"
            severity = IssueSeverity.WARNING
            self._repo.add_issue(
                upload,
                row_number=row_number,
                field=field,
                raw_value=raw_value,
                reason=full_reason,
                severity=severity,
            )
            return full_reason, severity

        self._repo.add_issue(
            upload,
            row_number=row_number,
            field=field,
            raw_value=raw_value,
            reason=reason,
        )
        return reason, IssueSeverity.ERROR

    def _prior_members_by_dodid(self) -> dict[str, MemberRecord]:
        active = self._repo.get_active()
        if active is None:
            return {}
        return {
            m.dodid: m
            for m in self._repo.list_members_for_upload(active.id)
        }

    @staticmethod
    def _from_record(record: MemberRecord) -> _AcceptedMember:
        return _AcceptedMember(
            dodid=record.dodid,
            display_name=record.display_name,
            rank=record.rank,
            personnel_type=record.personnel_type,
            afsc=record.afsc,
            normalized_afsc=record.normalized_afsc,
            source_row_number=record.source_row_number,
        )

    def list_uploads(self) -> list[UploadOut]:
        return [self._to_upload_out(upload) for upload in self._repo.list_all()]

    def _to_upload_out(self, upload: RosterUploadRecord) -> UploadOut:
        return UploadOut(
            upload_id=upload.id,
            filename=upload.filename,
            uploaded_at=upload.uploaded_at,
            status=upload.status,
            total_rows=upload.total_rows,
            accepted_rows=upload.accepted_rows,
            rejected_rows=upload.rejected_rows,
            ruleset_version=upload.ruleset_version,
            is_active=upload.is_active,
            issues=[
                IssueOut(
                    row=issue.row_number,
                    field=issue.field,
                    value=issue.raw_value,
                    reason=issue.reason,
                    severity=issue.severity,
                )
                for issue in upload.issues
            ],
        )

    def _to_preview_out(
        self,
        upload: RosterUploadRecord,
        successes: list[_RowResult],
        failures: list[_RowResult],
    ) -> PreviewOut:
        return PreviewOut(
            upload_id=upload.id,
            filename=upload.filename,
            ruleset_version=upload.ruleset_version,
            total_rows=upload.total_rows,
            accepted_rows=upload.accepted_rows,
            rejected_rows=upload.rejected_rows,
            can_commit=upload.status == UploadStatus.PENDING and upload.accepted_rows > 0,
            successes=[
                PreviewRowOut(
                    row=r.row,
                    status=r.status,
                    dodid=r.dodid,
                    display_name=r.display_name,
                    afsc=r.afsc,
                    normalized_afsc=r.normalized_afsc,
                    personnel_type=r.personnel_type,
                    reason=r.reason,
                    severity=r.severity,
                )
                for r in successes
            ],
            failures=[
                PreviewRowOut(
                    row=r.row,
                    status=r.status,
                    dodid=r.dodid,
                    display_name=r.display_name,
                    afsc=r.afsc,
                    normalized_afsc=r.normalized_afsc,
                    personnel_type=r.personnel_type,
                    reason=r.reason,
                    severity=r.severity,
                )
                for r in failures
            ],
        )

    def _parse_rows(self, content: bytes) -> list[dict[str, Any]]:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                raise RosterImportError("Workbook is empty")

            mapping = self._map_headers(header_row)
            if "dodid" not in mapping:
                raise RosterImportError(
                    "Workbook must include a DODID column "
                    f"(found: {[str(h) for h in header_row if h]})"
                )
            has_name = "display_name" in mapping or (
                "first_name" in mapping or "last_name" in mapping
            )
            if not has_name or "afsc" not in mapping:
                raise RosterImportError(
                    "Workbook must include name (or First/Last Name) and AFSC columns "
                    f"(found: {[str(h) for h in header_row if h]})"
                )

            results: list[dict[str, Any]] = []
            for offset, values in enumerate(rows_iter, start=2):
                if values is None or all(v is None or str(v).strip() == "" for v in values):
                    continue
                row: dict[str, Any] = {"row_number": offset}
                for field, idx in mapping.items():
                    raw = values[idx] if idx < len(values) else None
                    row[field] = self._cell_to_str(raw)
                if not row.get("display_name"):
                    parts = [p for p in (row.get("first_name"), row.get("last_name")) if p]
                    row["display_name"] = " ".join(parts) or None
                results.append(row)
            return results
        finally:
            wb.close()

    def _map_headers(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().lower().replace(" ", "_")
            for field, aliases in HEADER_ALIASES.items():
                if key in aliases and field not in mapping:
                    mapping[field] = idx
        return mapping

    @staticmethod
    def _cell_to_str(raw: Any) -> str | None:
        if raw is None:
            return None
        if isinstance(raw, bool):
            text = str(raw)
        elif isinstance(raw, int):
            text = str(raw)
        elif isinstance(raw, float):
            text = str(int(raw)) if raw.is_integer() else str(raw).rstrip("0").rstrip(".")
        else:
            text = str(raw).strip()
        text = text.strip()
        return text or None

    @staticmethod
    def _parse_personnel_type(value: str) -> PersonnelType | None:
        normalized = value.strip().lower()
        if normalized in {"enlisted", "e", "enl"}:
            return PersonnelType.ENLISTED
        if normalized in {"officer", "o", "off"}:
            return PersonnelType.OFFICER
        return None
